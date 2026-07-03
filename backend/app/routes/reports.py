from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
import io
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from app.db.db import db
from app.routes.dependencies import get_current_user, RoleChecker

router = APIRouter(prefix="/api/reports", tags=["reports"])

async def get_report_data(report_type: str) -> List[Dict[str, Any]]:
    if report_type == "inventory":
        return await db.inventory.find({}).to_list(1000)
    elif report_type == "requests":
        return await db.requests.find({}).to_list(1000)
    elif report_type == "purchases":
        return await db.purchase_orders.find({}).to_list(1000)
    elif report_type == "complaints":
        return await db.complaints.find({}).to_list(1000)
    elif report_type == "departments":
        return await db.departments.find({}).to_list(100)
    else:
        raise HTTPException(status_code=400, detail="Invalid report type")

# ----------------- CSV EXPORT -----------------
@router.get("/export/csv/{report_type}")
async def export_csv(report_type: str, current_user: dict = Depends(get_current_user)):
    data = await get_report_data(report_type)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    if not data:
        writer.writerow(["No data available"])
    else:
        # Generate headers based on keys, excluding internal password fields
        headers = [k for k in data[0].keys() if k != "hashed_password" and k != "reset_token"]
        writer.writerow(headers)
        
        for item in data:
            row = []
            for h in headers:
                val = item.get(h)
                if isinstance(val, list):
                    val = str(val)  # format lists
                row.append(val)
            writer.writerow(row)
            
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=sipms_{report_type}_report.csv"}
    )

# ----------------- EXCEL EXPORT -----------------
@router.get("/export/excel/{report_type}")
async def export_excel(report_type: str, current_user: dict = Depends(get_current_user)):
    data = await get_report_data(report_type)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{report_type.capitalize()} Report"
    
    # Premium Excel Styling
    title_font = Font(name="Arial", size=16, bold=True, color="1B365D")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Write Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Smart Inventory & Procurement Management System - {report_type.capitalize()} Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 40
    
    ws.append([]) # empty row
    
    if not data:
        ws.append(["No data available"])
    else:
        headers = [k for k in data[0].keys() if k != "hashed_password" and k != "reset_token"]
        ws.append(headers)
        
        # Style headers
        header_row_idx = 3
        ws.row_dimensions[header_row_idx].height = 25
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
        # Write data
        for item in data:
            row_data = []
            for h in headers:
                val = item.get(h)
                if isinstance(val, list) or isinstance(val, dict):
                    val = str(val)
                row_data.append(val)
            ws.append(row_data)
            
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.row < 3: continue
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sipms_{report_type}_report.xlsx"}
    )

# ----------------- PDF EXPORT -----------------
@router.get("/export/pdf/{report_type}")
async def export_pdf(report_type: str, current_user: dict = Depends(get_current_user)):
    data = await get_report_data(report_type)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name='TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#1B365D'),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        name='SubtitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#666666'),
        spaceAfter=25
    )
    cell_style = ParagraphStyle(
        name='CellStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10
    )
    
    # Document Header
    story.append(Paragraph(f"SIPMS {report_type.upper()} REPORT", title_style))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Target: Institutional Audit Dashboard", subtitle_style))
    
    if not data:
        story.append(Paragraph("No records found.", styles['Normal']))
    else:
        # Take up to top 6 keys to avoid horizontal overflow on letter paper
        all_keys = [k for k in data[0].keys() if k != "hashed_password" and k != "reset_token"]
        keys = all_keys[:6]
        
        # Headers Row
        table_data = [[Paragraph(f"<b>{k.replace('_', ' ').capitalize()}</b>", cell_style) for k in keys]]
        
        # Data Rows
        for item in data[:50]: # cap PDF preview to 50 items to keep file clean
            row = []
            for k in keys:
                val = item.get(k, "")
                if isinstance(val, list) or isinstance(val, dict):
                    val = str(val)[:30] # truncate nested fields
                else:
                    val = str(val)[:40]
                row.append(Paragraph(val, cell_style))
            table_data.append(row)
            
        # Calculate widths dynamically
        col_width = 540 / len(keys)
        t = Table(table_data, colWidths=[col_width] * len(keys))
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('TOPPADDING', (0,0), (-1,0), 6),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#F8FAFC'), colors.white]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        
        story.append(t)
        
    doc.build(story)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=sipms_{report_type}_report.pdf"}
    )
